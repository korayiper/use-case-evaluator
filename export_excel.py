"""Build .xlsx workbooks for download via the /api/export/* routes in
app.py - the counterpart to import_excel.py's CLI-driven Excel -> app
import.
"""

import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PRIORITIZED_COLUMNS = [
    ("Name", "name"),
    ("Ideengeber", "idea_initiator"),
    ("Nutzungskategorie", "use_category_label"),
    ("Wirtschaftlicher Nutzen", "economic_value_label"),
    ("Priorität", "priority"),
    ("Entwicklungsdauer", "development_time_label"),
    ("Go-Live", "golive_date"),
    ("Priorisiert", "prioritized_round"),
]

BOARD_COLUMNS = [
    "Name",
    "Nutzungskategorie",
    "Entwicklungsdauer",
    "Priorität",
    "Wirtschaftlicher Nutzen",
    "Wichtig",
]


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in rows:
        ws.append(row)
    # Rough auto-width: widest of the header or any cell in that column,
    # capped so one long name doesn't blow out the whole sheet.
    for i, header in enumerate(headers, start=1):
        widest = max([len(header)] + [len(str(row[i - 1])) for row in rows], default=len(header))
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(45, widest + 2))


def build_prioritized_workbook(use_cases: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Priorisiert"
    headers = [label for label, _ in PRIORITIZED_COLUMNS]
    rows = [[uc.get(key) or "" for _, key in PRIORITIZED_COLUMNS] for uc in use_cases]
    _write_sheet(ws, headers, rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_board_workbook(board: list[dict]) -> io.BytesIO:
    """Exactly what's on the board page, in its current order - same
    columns as the on-screen table (minus the Aktionen column, which has
    nothing to export)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Board"
    headers = ["#", *BOARD_COLUMNS]
    rows = [
        [
            i,
            uc["name"],
            uc["use_category_label"],
            uc["development_time_label"],
            uc["priority"],
            uc["economic_value_label"],
            f"Wichtig ({len(uc['important_departments'])}/6)" if uc["is_important"] else "–",
        ]
        for i, uc in enumerate(board, start=1)
    ]
    _write_sheet(ws, headers, rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
