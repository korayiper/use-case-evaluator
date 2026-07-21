"""Import AI use cases from an Excel sheet into the running app, via its API.

Usage:
    uv run python import_excel.py --file path/to/sheet.xlsx

By default this only validates the sheet and prints a report - nothing is
written. Pass --commit to actually create the use cases:

    uv run python import_excel.py --file path/to/sheet.xlsx --commit

Column matching is alias-based (see FIELD_ALIASES below) and
case-/whitespace-insensitive, so exact header wording doesn't have to match
perfectly. Always run once WITHOUT --commit first and check the "Spalten-
zuordnung" (column mapping) section - if a column isn't picked up the way you
expect, add an alias for it below and run again.

Columns from the sheet that aren't part of the app's data model (e.g.
"Referenz", "Quelle bzw. Datenart", "Backlog Notizen", "Priorisierung") are
read but never imported - priority is always computed by the app from the
four scored attributes, never taken from the sheet.

Re-running this script on the same rows creates duplicate use cases - there
is currently no identifier in the app's data model to match sheet rows back
to existing entries.
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests

import scoring

# Each field maps to a list of alias substrings to look for in a (lowercased,
# whitespace-normalized) column header. When several aliases match a header,
# the longest one wins - so e.g. "Beschreibung Nutzen" resolves to
# value_added_description rather than the bare "nutzen" alias for value_added.
FIELD_ALIASES = {
    "name": ["name"],
    "idea_initiator": ["idee-initiator", "idee initiator", "ideeninitiator", "ideengeber"],
    "description": [
        "beschreibung use case",
        "beschreibung des use case",
        "beschreibung anwendungsfall",
        "beschreibung des anwendungsfalls",
    ],
    "value_added_description": ["beschreibung nutzen", "beschreibung des nutzens", "beschreibung mehrwert"],
    "use_category": ["nutzungskategorie", "kategorie"],
    "value_added": ["nutzen"],
    "development_time": ["entwicklungszeit"],
    "process_criticality": ["prozesskritikalität", "prozesskritikalitaet"],
    "process_dependency": ["prozessintegration", "prozessabhängigkeit", "abhängigkeit"],
    "ai_feasibility": ["machbarkeit"],
    "golive_date": ["einführungszeitpunkt", "einfuehrungszeitpunkt", "go-live", "golive"],
}

REQUIRED_FIELDS = [
    "name",
    "idea_initiator",
    "use_category",
    "ai_feasibility",
    "value_added",
    "development_time",
    "process_criticality",
    "process_dependency",
    "golive_date",
]
OPTIONAL_FIELDS = ["description", "value_added_description"]

# value -> German label dicts from scoring.py, reversed to label -> value, so
# an Excel cell's German text can be translated back to the app's internal
# keys. Single source of truth stays in scoring.py.
VALUE_MAPS = {
    "use_category": {label.strip().lower(): value for value, label in scoring.USE_CATEGORY.items()},
    "value_added": {label.strip().lower(): value for value, label in scoring.VALUE_ADDED_LABELS.items()},
    "development_time": {
        label.strip().lower(): value for value, label in scoring.DEVELOPMENT_TIME_LABELS.items()
    },
    "process_criticality": {
        label.strip().lower(): value for value, label in scoring.PROCESS_CRITICALITY_LABELS.items()
    },
    "process_dependency": {
        label.strip().lower(): value for value, label in scoring.PROCESS_DEPENDENCY_LABELS.items()
    },
    "ai_feasibility": {label.strip().lower(): value for value, label in scoring.AI_FEASIBILITY_LABELS.items()},
}

DATE_FORMATS = ["%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y"]


def normalize(text) -> str:
    return " ".join(str(text).strip().lower().split())


def match_column(header) -> str | None:
    normalized = normalize(header)
    best_field, best_len = None, -1
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized and len(alias) > best_len:
                best_field, best_len = field, len(alias)
    return best_field


def parse_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_value(field: str, raw) -> tuple[str | None, str | None]:
    """Returns (parsed_value, error) - error is None on success."""
    if raw is None or str(raw).strip() == "":
        return None, "leer"
    if field == "golive_date":
        parsed = parse_date(raw)
        if parsed is None:
            return None, f"Datum nicht erkannt: {raw!r}"
        return parsed, None
    if field in VALUE_MAPS:
        mapped = VALUE_MAPS[field].get(normalize(raw))
        if mapped is None:
            allowed = ", ".join(sorted(set(VALUE_MAPS[field].keys())))
            return None, f"unbekannter Wert {raw!r} (erwartet: {allowed})"
        return mapped, None
    return str(raw).strip(), None


def build_column_map(headers) -> tuple[dict[int, str], list[str]]:
    column_map, unmapped = {}, []
    for idx, header in enumerate(headers):
        if header is None or str(header).strip() == "":
            continue
        field = match_column(header)
        if field:
            column_map[idx] = field
        else:
            unmapped.append(str(header))
    return column_map, unmapped


def build_payload(column_map: dict[int, str], row) -> tuple[dict, list[str]]:
    payload = {field: "" for field in OPTIONAL_FIELDS}
    errors = []
    for idx, field in column_map.items():
        raw = row[idx] if idx < len(row) else None
        value, error = parse_value(field, raw)
        if error:
            if field in REQUIRED_FIELDS:
                errors.append(f"{field}: {error}")
            continue
        payload[field] = value
    return payload, errors


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--file", required=True, type=Path, help="Pfad zur Excel-Datei (.xlsx)")
    parser.add_argument("--sheet", default=None, help="Blattname (Standard: aktives Blatt)")
    parser.add_argument("--api-base", default="http://127.0.0.1:8010", help="Basis-URL der laufenden App")
    parser.add_argument("--commit", action="store_true", help="Tatsächlich importieren (sonst nur Vorschau)")
    args = parser.parse_args()

    if not args.file.exists():
        print(f"Datei nicht gefunden: {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    column_map, unmapped = build_column_map(headers)

    print("Spaltenzuordnung:")
    for idx, header in enumerate(headers):
        if header is None or str(header).strip() == "":
            continue
        print(f"  {str(header)!r:45s} -> {column_map.get(idx) or '(ignoriert)'}")
    print()

    missing_required = [f for f in REQUIRED_FIELDS if f not in column_map.values()]
    if missing_required:
        print("Für folgende Pflichtfelder wurde keine Spalte gefunden:")
        for field in missing_required:
            print(f"  - {field}")
        print("Bitte FIELD_ALIASES in diesem Skript ergänzen und erneut versuchen.")
        sys.exit(1)

    valid_rows, invalid_rows = [], []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        payload, errors = build_payload(column_map, row)
        label = payload.get("name") or f"Zeile {row_num}"
        if errors:
            invalid_rows.append((row_num, label, errors))
        else:
            valid_rows.append((row_num, label, payload))

    print(f"{len(valid_rows)} gültige Zeile(n), {len(invalid_rows)} mit Fehlern.\n")

    if invalid_rows:
        print("Fehlerhafte Zeilen (werden nicht importiert):")
        for row_num, label, errors in invalid_rows:
            print(f"  Zeile {row_num} ({label}):")
            for error in errors:
                print(f"    - {error}")
        print()

    if not args.commit:
        print("Vorschau-Modus (kein --commit) - es wurde nichts geschrieben.")
        if valid_rows:
            print("\nBeispiel (erste gültige Zeile):")
            for key, value in valid_rows[0][2].items():
                print(f"  {key}: {value!r}")
        return

    created, failed = 0, 0
    for row_num, label, payload in valid_rows:
        resp = requests.post(f"{args.api_base}/api/use-cases", json=payload, timeout=10)
        if resp.ok:
            created += 1
        else:
            failed += 1
            print(f"  Zeile {row_num} ({label}) fehlgeschlagen: {resp.status_code} {resp.text}")

    print(f"\n{created} Anwendungsfälle importiert, {failed} fehlgeschlagen.")


if __name__ == "__main__":
    main()
